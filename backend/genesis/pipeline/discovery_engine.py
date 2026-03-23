"""Discovery Engine — the brain behind the guided build conversation.

Real website scanning, file analysis (including Excel/images), and synthesis.
"""

from __future__ import annotations

import io
import json
import logging
import re
from typing import Any
from urllib.parse import urlparse

import httpx

from genesis.config import settings

logger = logging.getLogger(__name__)


# ── Website Scanner ───────────────────────────────────────────────────────────


async def scan_website(url: str) -> dict[str, Any]:
    """Fetch a website and extract meaningful information."""
    result: dict[str, Any] = {
        "url": url,
        "hostname": urlparse(url).hostname or "",
        "status": "error",
    }

    try:
        async with httpx.AsyncClient(
            follow_redirects=True,
            timeout=15.0,
            headers={"User-Agent": "Genesis SaaS Scanner/1.0"},
        ) as client:
            resp = await client.get(url)
            result["status_code"] = resp.status_code
            html = resp.text

            if resp.status_code != 200:
                result["status"] = "error"
                result["error"] = f"HTTP {resp.status_code}"
                return result

            result["status"] = "success"

            title_match = re.search(r"<title[^>]*>(.*?)</title>", html, re.IGNORECASE | re.DOTALL)
            result["title"] = title_match.group(1).strip() if title_match else ""

            desc_match = re.search(
                r'<meta[^>]*name=["\']description["\'][^>]*content=["\'](.*?)["\']',
                html, re.IGNORECASE,
            )
            result["description"] = desc_match.group(1).strip() if desc_match else ""

            tech = []
            tech_patterns = {
                "React": r"react|__next|_next/static|reactroot",
                "Next.js": r"__next|_next/static|next/dist",
                "Angular": r"ng-version|angular\.js|ng-app|zone\.js",
                "Vue.js": r"vue\.js|vuejs|__vue|v-app",
                "Tailwind CSS": r"tailwind|tw-",
                "Bootstrap": r"bootstrap\.css|bootstrap\.min",
                "WordPress": r"wp-content|wp-includes",
                "Shopify": r"shopify|cdn\.shopify",
                "Stripe": r"stripe\.js|js\.stripe\.com",
                "Google Analytics": r"google-analytics|gtag|ga\.js",
            }
            for name, pattern in tech_patterns.items():
                if re.search(pattern, html, re.IGNORECASE):
                    tech.append(name)
            result["tech_stack"] = tech

            nav_links = re.findall(
                r'<a[^>]*href=["\'](/[^"\'#]*)["\'][^>]*>(.*?)</a>',
                html, re.IGNORECASE | re.DOTALL,
            )
            features = []
            for href, text in nav_links:
                clean = re.sub(r"<[^>]+>", "", text).strip()
                if clean and len(clean) < 50 and href.count("/") <= 2:
                    features.append({"path": href, "label": clean})
            result["navigation"] = features[:20]

            headings = []
            for tag in ["h1", "h2"]:
                for match in re.finditer(
                    rf"<{tag}[^>]*>(.*?)</{tag}>", html, re.IGNORECASE | re.DOTALL
                ):
                    clean = re.sub(r"<[^>]+>", "", match.group(1)).strip()
                    if clean and len(clean) < 200:
                        headings.append(clean)
            result["headings"] = headings[:10]

            colors = set(re.findall(r"#[0-9a-fA-F]{6}", html))
            result["colors"] = sorted(colors)[:10]
            result["content_length"] = len(html)
            result["has_forms"] = bool(re.search(r"<form", html, re.IGNORECASE))
            result["has_login"] = bool(re.search(r"login|sign.?in|log.?in", html, re.IGNORECASE))
            result["has_pricing"] = bool(re.search(r"pricing|plans|subscribe", html, re.IGNORECASE))
            result["has_api_docs"] = bool(re.search(r"api|developer|documentation", html, re.IGNORECASE))

            summary_parts = []
            if result["title"]:
                summary_parts.append(f"**{result['title']}**")
            if result["description"]:
                summary_parts.append(result["description"])
            if tech:
                summary_parts.append(f"Tech: {', '.join(tech)}")
            if result["has_pricing"]:
                summary_parts.append("Has pricing page")
            if result["has_login"]:
                summary_parts.append("Has login/auth")
            if headings:
                summary_parts.append(f"Key messages: {'; '.join(headings[:3])}")
            result["summary"] = " | ".join(summary_parts)

    except httpx.TimeoutException:
        result["status"] = "error"
        result["error"] = "Request timed out (15s)"
    except Exception as e:
        result["status"] = "error"
        result["error"] = str(e)[:200]

    return result


# ── File Analyzer ─────────────────────────────────────────────────────────────


def analyze_file_text(content: str, filename: str, file_type: str) -> dict[str, Any]:
    """Analyze text-based file content."""
    result: dict[str, Any] = {
        "filename": filename,
        "type": file_type,
        "size": len(content),
    }

    if file_type in ("text/csv", "application/csv") or filename.endswith(".csv"):
        lines = content.strip().split("\n")
        if lines:
            headers = [h.strip().strip('"') for h in lines[0].split(",")]
            result["format"] = "csv"
            result["columns"] = headers
            result["row_count"] = len(lines) - 1
            result["summary"] = f"CSV with {len(headers)} columns, {len(lines)-1} rows: {', '.join(headers[:8])}"
            if len(headers) > 8:
                result["summary"] += f" (+{len(headers)-8} more)"
            # Sample data
            if len(lines) > 1:
                result["sample_rows"] = lines[1:min(4, len(lines))]

    elif file_type == "application/json" or filename.endswith(".json"):
        try:
            data = json.loads(content)
            if isinstance(data, list):
                result["format"] = "json_array"
                result["item_count"] = len(data)
                if data and isinstance(data[0], dict):
                    result["fields"] = list(data[0].keys())
                result["summary"] = f"JSON array with {len(data)} items"
            elif isinstance(data, dict):
                result["format"] = "json_object"
                result["keys"] = list(data.keys())[:20]
                result["summary"] = f"JSON object with keys: {', '.join(list(data.keys())[:10])}"
        except json.JSONDecodeError:
            result["summary"] = "Invalid JSON file"

    elif file_type in ("text/plain", "text/markdown") or filename.endswith((".txt", ".md")):
        lines = content.split("\n")
        words = len(content.split())
        headings = [l.strip() for l in lines if l.strip().startswith("#")]
        result["format"] = "text"
        result["word_count"] = words
        result["line_count"] = len(lines)
        result["headings"] = headings[:10]
        result["summary"] = f"{'Markdown' if filename.endswith('.md') else 'Text'}: {words} words, {len(lines)} lines"
        if headings:
            result["summary"] += f". Sections: {', '.join(h.lstrip('#').strip() for h in headings[:5])}"
        result["preview"] = content[:1000]

    else:
        result["format"] = "text"
        result["summary"] = f"Text file: {filename} ({len(content)} chars)"
        result["preview"] = content[:1000]

    return result


def analyze_file_binary(content: bytes, filename: str, file_type: str) -> dict[str, Any]:
    """Analyze binary file content (Excel, images, etc.)."""
    result: dict[str, Any] = {
        "filename": filename,
        "type": file_type,
        "size": len(content),
    }

    # Excel files
    if filename.endswith((".xlsx", ".xls")) or file_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheets = wb.sheetnames
            result["format"] = "excel"
            result["sheets"] = sheets

            all_data = []
            for sheet_name in sheets[:3]:  # Max 3 sheets
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(max_row=min(ws.max_row or 1, 50), values_only=True))
                if rows:
                    headers = [str(c) if c else f"col_{i}" for i, c in enumerate(rows[0])]
                    data_rows = rows[1:min(6, len(rows))]  # Sample 5 rows
                    all_data.append({
                        "sheet": sheet_name,
                        "columns": headers,
                        "row_count": (ws.max_row or 1) - 1,
                        "col_count": len(headers),
                        "sample": [
                            {h: str(v) if v is not None else "" for h, v in zip(headers, row)}
                            for row in data_rows
                        ],
                    })
            wb.close()

            result["sheet_data"] = all_data
            # Build summary
            total_rows = sum(s["row_count"] for s in all_data)
            total_cols = sum(s["col_count"] for s in all_data)
            result["columns"] = all_data[0]["columns"] if all_data else []
            result["row_count"] = total_rows

            sheet_summaries = []
            for s in all_data:
                sheet_summaries.append(
                    f"Sheet '{s['sheet']}': {s['row_count']} rows, {s['col_count']} columns ({', '.join(s['columns'][:6])})"
                )
            result["summary"] = f"Excel workbook with {len(sheets)} sheet(s), {total_rows} total rows. " + "; ".join(sheet_summaries)

            # Build a text preview for Claude
            preview_parts = []
            for s in all_data:
                preview_parts.append(f"\n--- Sheet: {s['sheet']} ({s['row_count']} rows) ---")
                preview_parts.append("Columns: " + ", ".join(s["columns"]))
                for row in s.get("sample", [])[:3]:
                    preview_parts.append(str(row))
            result["preview"] = "\n".join(preview_parts)

        except Exception as e:
            result["format"] = "excel"
            result["summary"] = f"Excel file: {filename} (failed to read: {str(e)[:100]})"
            logger.warning("Excel parse error: %s", e)

    # Images
    elif file_type.startswith("image/"):
        result["format"] = "image"
        img_type = file_type.split("/")[-1].upper()
        result["summary"] = f"Image ({img_type}, {len(content):,} bytes)"
        # Store base64 so Claude can analyze it later
        import base64
        result["base64"] = base64.b64encode(content).decode()
        result["preview"] = f"[Image: {filename} — {img_type} format, {len(content):,} bytes. Claude can analyze this image for UI/UX reference.]"

    # PDF
    elif filename.endswith(".pdf") or file_type == "application/pdf":
        result["format"] = "pdf"
        result["summary"] = f"PDF document: {filename} ({len(content):,} bytes)"
        result["preview"] = "[PDF document — text extraction requires additional processing]"

    else:
        result["format"] = "binary"
        result["summary"] = f"Binary file: {filename} ({file_type}, {len(content):,} bytes)"

    return result


# Legacy wrapper
def analyze_file(content: str, filename: str, file_type: str) -> dict[str, Any]:
    """Analyze text file content (backward compatible)."""
    return analyze_file_text(content, filename, file_type)


# ── Synthesis Engine ──────────────────────────────────────────────────────────


def synthesize_discovery(
    messages: list[dict],
    context: dict[str, Any],
) -> dict[str, Any]:
    """Synthesize all discovery data into structured artifacts."""
    user_messages = [m["content"] for m in messages if m.get("role") == "user"]
    assistant_messages = [m["content"] for m in messages if m.get("role") == "assistant"]
    all_text = " ".join(user_messages + assistant_messages).lower()

    uploads = context.get("uploads", [])
    scans = context.get("scans", [])

    artifacts: dict[str, Any] = {}

    # ── Extract Personas ──
    personas = []
    persona_keywords = [
        (r"(?:admin|administrator)", "Administrator"),
        (r"(?:doctor|physician|clinician)", "Healthcare Provider"),
        (r"(?:patient|client|customer)", "Patient/Client"),
        (r"(?:manager|supervisor|lead)", "Manager"),
        (r"(?:user|end.?user|member)", "End User"),
        (r"(?:developer|engineer|dev)", "Developer"),
        (r"(?:owner|business.?owner|entrepreneur)", "Business Owner"),
        (r"(?:staff|employee|team.?member|worker)", "Staff Member"),
        (r"(?:vendor|supplier|partner)", "Vendor/Partner"),
        (r"(?:receptionist|front.?desk|secretary)", "Receptionist"),
        (r"(?:florist|designer|artist)", "Florist/Designer"),
        (r"(?:bride|groom|couple|wedding)", "Wedding Client"),
    ]
    for pattern, label in persona_keywords:
        if re.search(pattern, all_text):
            for msg in user_messages:
                match = re.search(rf"(.{{0,100}}{pattern}.{{0,100}})", msg.lower())
                if match:
                    personas.append({"role": label, "context": match.group(1).strip()})
                    break
            else:
                personas.append({"role": label, "context": ""})
    artifacts["personas"] = personas

    # ── Extract Features ──
    features = []
    feature_patterns = [
        r"(?:need|want|should|must|require)\s+(?:to\s+)?(?:be able to\s+)?(.{10,80}?)(?:\.|,|$)",
        r"(?:feature|functionality|capability)[:\s]+(.{10,80}?)(?:\.|,|$)",
        r"(?:ability to|able to)\s+(.{10,80}?)(?:\.|,|$)",
    ]
    seen = set()
    for pattern in feature_patterns:
        for msg in user_messages:
            for match in re.finditer(pattern, msg, re.IGNORECASE):
                feat = match.group(1).strip().rstrip(".,;")
                if feat.lower() not in seen and len(feat) > 10:
                    features.append(feat)
                    seen.add(feat.lower())
    artifacts["features"] = features[:15]

    # ── Extract Constraints ──
    constraints = []
    constraint_patterns = [
        (r"(?:hipaa|phi|protected health)", "HIPAA compliance required"),
        (r"(?:gdpr|data protection|privacy)", "GDPR/data privacy compliance"),
        (r"(?:pci|payment|credit card)", "PCI-DSS payment security"),
        (r"(?:soc.?2|audit)", "SOC 2 compliance"),
        (r"(?:budget|cost|cheap|affordable|free tier)", "Budget constraints"),
        (r"(?:deadline|timeline|by next|asap|urgent)", "Timeline pressure"),
        (r"(?:mobile|responsive|phone|tablet)", "Mobile/responsive required"),
        (r"(?:integrat|connect|sync|api|third.?party)", "Integration requirements"),
        (r"(?:offline|no internet|disconnected)", "Offline capability needed"),
        (r"(?:multi.?tenant|saas|multiple.?org)", "Multi-tenant architecture"),
    ]
    for pattern, constraint in constraint_patterns:
        if re.search(pattern, all_text):
            constraints.append(constraint)
    artifacts["constraints"] = constraints

    # ── Extract Problems ──
    problems = []
    problem_patterns = [
        r"(?:problem|issue|challenge|pain.?point|frustrat|difficult|struggle)[:\s]+(.{10,100}?)(?:\.|$)",
        r"(?:currently|right now|today)[,\s]+(.{10,100}?)(?:\.|$)",
        r"(?:manual|tedious|slow|error.?prone|unreliable)(.{0,80}?)(?:\.|$)",
    ]
    for pattern in problem_patterns:
        for msg in user_messages:
            for match in re.finditer(pattern, msg, re.IGNORECASE):
                prob = match.group(1).strip().rstrip(".,;")
                if len(prob) > 10:
                    problems.append(prob)
    artifacts["problems"] = problems[:8]

    # ── Progress ──
    progress = {
        "personas_identified": len(personas) > 0,
        "problems_defined": len(problems) > 0,
        "features_explored": len(features) > 0,
        "constraints_identified": len(constraints) > 0,
        "references_shared": len(uploads) + len(scans) > 0,
        "deep_conversation": len(messages) >= 8,
        "scope_discussed": bool(re.search(r"mvp|scope|first version|priority|must.?have|v1", all_text)),
        "message_count": len(messages),
        "user_message_count": len(user_messages),
    }
    checks = [v for k, v in progress.items() if isinstance(v, bool)]
    progress["percent"] = round((sum(checks) / max(1, len(checks))) * 100)
    artifacts["progress"] = progress

    # ── Reference Materials ──
    artifacts["scans"] = [
        {
            "url": s.get("url", ""),
            "title": s.get("title", ""),
            "summary": s.get("summary", ""),
            "tech_stack": s.get("tech_stack", []),
            "has_login": s.get("has_login", False),
            "has_pricing": s.get("has_pricing", False),
        }
        for s in scans
    ]
    artifacts["uploads"] = [
        {
            "name": u.get("name", ""),
            "type": u.get("type", ""),
            "summary": u.get("summary", ""),
            "columns": u.get("columns"),
            "row_count": u.get("row_count"),
            "preview": u.get("preview", "")[:300],
            "format": u.get("format", ""),
        }
        for u in uploads
    ]

    return artifacts
